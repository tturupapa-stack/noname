import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Load data
with open('../backend/data/briefings.json', 'r') as f:
    data = json.load(f)

wb = Workbook()

# ========== Sheet 1: Daily Hot Stocks ==========
ws1 = wb.active
ws1.title = "Daily Hot Stocks"

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1a1a2e")
positive_font = Font(color="00AA00")
negative_font = Font(color="DD0000")
blue_font = Font(color="0000FF")
border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
center = Alignment(horizontal='center', vertical='center')
yellow_fill = PatternFill("solid", fgColor="FFFF00")

# Headers
headers = ["Date", "Symbol", "Company", "Price ($)", "Change (%)", "Volume", "Avg Volume", "Vol Ratio", "Market Cap ($B)", "Total Score"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# Data rows
for row_idx, item in enumerate(data, 2):
    stock = item['stock']
    score = item['score']

    ws1.cell(row=row_idx, column=1, value=item['date']).border = border
    ws1.cell(row=row_idx, column=2, value=stock['symbol']).border = border
    ws1.cell(row=row_idx, column=2).font = Font(bold=True)
    ws1.cell(row=row_idx, column=3, value=stock['name']).border = border
    ws1.cell(row=row_idx, column=4, value=stock['price']).border = border
    ws1.cell(row=row_idx, column=4).number_format = '$#,##0.00'

    change_cell = ws1.cell(row=row_idx, column=5, value=stock['change_percent'])
    change_cell.border = border
    change_cell.number_format = '0.00%'
    change_cell.font = positive_font if stock['change_percent'] > 0 else negative_font

    ws1.cell(row=row_idx, column=6, value=stock['volume']).border = border
    ws1.cell(row=row_idx, column=6).number_format = '#,##0'
    ws1.cell(row=row_idx, column=7, value=stock['avg_volume']).border = border
    ws1.cell(row=row_idx, column=7).number_format = '#,##0'
    ws1.cell(row=row_idx, column=8, value=stock['volume_ratio']).border = border
    ws1.cell(row=row_idx, column=8).number_format = '0.00'

    mcap = stock['market_cap'] / 1e9 if stock['market_cap'] else 0
    ws1.cell(row=row_idx, column=9, value=mcap).border = border
    ws1.cell(row=row_idx, column=9).number_format = '$#,##0.00'

    ws1.cell(row=row_idx, column=10, value=score['total']).border = border
    ws1.cell(row=row_idx, column=10).alignment = center

# Summary row with formulas
last_row = len(data) + 1
summary_row = last_row + 2

ws1.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
ws1.cell(row=summary_row, column=4, value="Avg Price")
ws1.cell(row=summary_row+1, column=4, value=f"=AVERAGE(D2:D{last_row})")
ws1.cell(row=summary_row+1, column=4).number_format = '$#,##0.00'
ws1.cell(row=summary_row+1, column=4).font = blue_font

ws1.cell(row=summary_row, column=5, value="Avg Change")
ws1.cell(row=summary_row+1, column=5, value=f"=AVERAGE(E2:E{last_row})")
ws1.cell(row=summary_row+1, column=5).number_format = '0.00%'
ws1.cell(row=summary_row+1, column=5).font = blue_font

ws1.cell(row=summary_row, column=8, value="Avg Vol Ratio")
ws1.cell(row=summary_row+1, column=8, value=f"=AVERAGE(H2:H{last_row})")
ws1.cell(row=summary_row+1, column=8).number_format = '0.00'
ws1.cell(row=summary_row+1, column=8).font = blue_font

ws1.cell(row=summary_row, column=10, value="Avg Score")
ws1.cell(row=summary_row+1, column=10, value=f"=AVERAGE(J2:J{last_row})")
ws1.cell(row=summary_row+1, column=10).number_format = '0.0'
ws1.cell(row=summary_row+1, column=10).font = blue_font

# Column widths
col_widths = [12, 10, 30, 12, 12, 15, 15, 12, 15, 12]
for i, width in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = width

# ========== Sheet 2: Score Analysis ==========
ws2 = wb.create_sheet("Score Analysis")

headers2 = ["Date", "Symbol", "Volume Score", "Price Change Score", "Momentum Score", "Market Cap Score", "Total Score"]
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

for row_idx, item in enumerate(data, 2):
    score = item['score']
    ws2.cell(row=row_idx, column=1, value=item['date']).border = border
    ws2.cell(row=row_idx, column=2, value=item['stock']['symbol']).border = border
    ws2.cell(row=row_idx, column=2).font = Font(bold=True)
    ws2.cell(row=row_idx, column=3, value=score['volume_score']).border = border
    ws2.cell(row=row_idx, column=4, value=score['price_change_score']).border = border
    ws2.cell(row=row_idx, column=5, value=score['momentum_score']).border = border
    ws2.cell(row=row_idx, column=6, value=score['market_cap_score']).border = border

    # Total as formula
    ws2.cell(row=row_idx, column=7, value=f"=SUM(C{row_idx}:F{row_idx})")
    ws2.cell(row=row_idx, column=7).border = border
    ws2.cell(row=row_idx, column=7).font = blue_font

# Score averages
avg_row = len(data) + 3
ws2.cell(row=avg_row, column=2, value="Average").font = Font(bold=True)
for col in range(3, 8):
    ws2.cell(row=avg_row, column=col, value=f"=AVERAGE({get_column_letter(col)}2:{get_column_letter(col)}{len(data)+1})")
    ws2.cell(row=avg_row, column=col).font = blue_font
    ws2.cell(row=avg_row, column=col).number_format = '0.0'

for i, width in enumerate([12, 10, 14, 18, 16, 18, 12], 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

# ========== Sheet 3: Stock Summary ==========
ws3 = wb.create_sheet("Stock Summary")

# Count unique stocks
from collections import defaultdict
stock_data = defaultdict(list)
for item in data:
    symbol = item['stock']['symbol']
    stock_data[symbol].append(item)

headers3 = ["Symbol", "Company", "Appearances", "Avg Price ($)", "Avg Change (%)", "Max Vol Ratio", "Avg Score", "Best Date"]
for col, header in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

row = 2
for symbol, items in stock_data.items():
    ws3.cell(row=row, column=1, value=symbol).border = border
    ws3.cell(row=row, column=1).font = Font(bold=True)
    ws3.cell(row=row, column=2, value=items[0]['stock']['name']).border = border
    ws3.cell(row=row, column=3, value=len(items)).border = border
    ws3.cell(row=row, column=3).alignment = center

    avg_price = sum(i['stock']['price'] for i in items) / len(items)
    ws3.cell(row=row, column=4, value=avg_price).border = border
    ws3.cell(row=row, column=4).number_format = '$#,##0.00'

    avg_change = sum(i['stock']['change_percent'] for i in items) / len(items)
    change_cell = ws3.cell(row=row, column=5, value=avg_change/100)
    change_cell.border = border
    change_cell.number_format = '0.00%'
    change_cell.font = positive_font if avg_change > 0 else negative_font

    max_vol = max(i['stock']['volume_ratio'] for i in items)
    vol_cell = ws3.cell(row=row, column=6, value=max_vol)
    vol_cell.border = border
    vol_cell.number_format = '0.00'
    if max_vol > 10:
        vol_cell.fill = yellow_fill

    avg_score = sum(i['score']['total'] for i in items) / len(items)
    ws3.cell(row=row, column=7, value=avg_score).border = border
    ws3.cell(row=row, column=7).number_format = '0.0'

    best = max(items, key=lambda x: x['score']['total'])
    ws3.cell(row=row, column=8, value=best['date']).border = border

    row += 1

for i, width in enumerate([10, 30, 14, 14, 14, 14, 12, 12], 1):
    ws3.column_dimensions[get_column_letter(i)].width = width

# ========== Sheet 4: Why Hot Analysis ==========
ws4 = wb.create_sheet("Why Hot Analysis")

headers4 = ["Date", "Symbol", "Reason 1", "Reason 2", "Reason 3", "Reason 4"]
for col, header in enumerate(headers4, 1):
    cell = ws4.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

for row_idx, item in enumerate(data, 2):
    ws4.cell(row=row_idx, column=1, value=item['date']).border = border
    ws4.cell(row=row_idx, column=2, value=item['stock']['symbol']).border = border
    ws4.cell(row=row_idx, column=2).font = Font(bold=True)

    for i, reason in enumerate(item.get('why_hot', [])[:4], 3):
        ws4.cell(row=row_idx, column=i, value=reason['message']).border = border

for i, width in enumerate([12, 10, 35, 35, 35, 35], 1):
    ws4.column_dimensions[get_column_letter(i)].width = width

# Save
output_path = 'hot-stocks-weekly-analysis.xlsx'
wb.save(output_path)
print(f"Excel file saved: {output_path}")
